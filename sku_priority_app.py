import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
            
st.title(" SKU Priority Highlighter (Excel)")
                
# Loading master priority list
priority_df = pd.read_excel("Final Priority List.xlsx")
priority_df = pd.read_excel(PRIORITY_FILE)
priority_skus = set(priority_df.iloc[:, 0].astype(str).str.strip())
    
# Uploading indent file
uploaded_file = st.file_uploader("Upload Loading Indent Excel", type=["xlsx"])
if uploaded_file:
    df = pd.read_excel(uploaded_file)
    df["Material"] = df["Material"].astype(str).str.strip()
                       
    # Loading Excel in openpyxl
    wb = load_workbook(uploaded_file)
    ws = wb.active
    
    # Defining highlighter style
    highlight_font = Font(color="FF0000", bold=True)  # Red bold text
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill
    
    # Applying formatting to matching rows (skip header, start from row 2)
    for i, row in enumerate(df.itertuples(index=False), start=2):
        material = str(row.Material).strip()
        if material in priority_skus:
            for col in range(1, 4):  # Applying to Material, Desc, QTY
                cell = ws.cell(row=i, column=col)
                cell.font = highlight_font
                cell.fill = highlight_fill

    # Saving updated workbook to buffer   
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    st.success(" Excel updated with priority highlights.")
    st.download_button("Download Highlighted Excel", data=output,
                       file_name="Highlighted_Indent.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
